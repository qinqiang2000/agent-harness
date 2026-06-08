"""
Excel 过滤处理脚本 - 删除不需要的出库单行
过滤规则从 config.json 中读取，无需改代码即可调整过滤条件。

用法:
  python excel_filter.py --file <excel文件路径>
"""

import argparse
import json
import sys
from pathlib import Path


def _load_filter_config() -> dict:
    """
    从 config.json 读取过滤规则。
    返回包含 excluded_materials 和 exclude_contract_types 的字典。
    """
    config_path = Path(__file__).parent / "config.json"
    if not config_path.exists():
        print("⚠️  config.json 不存在，使用空过滤规则", file=sys.stderr)
        return {"excluded_materials": [], "exclude_contract_types": []}
    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)
    return config.get(
        "excel_filter", {"excluded_materials": [], "exclude_contract_types": []}
    )


def filter_excel(file_path: str) -> bool:
    """
    过滤 Excel 文件：
    1. 删除"合同业务类型"包含 exclude_contract_types 任意关键词的行
    2. 删除"物料名称"精确匹配 excluded_materials 的行
    过滤规则从 config.json 读取，支持动态调整。

    Args:
        file_path: Excel 文件路径

    Returns:
        True 表示处理成功
    """
    try:
        import openpyxl
    except ImportError:
        print("❌ 缺少依赖 openpyxl，请运行 pip install openpyxl", file=sys.stderr)
        return False

    filter_config = _load_filter_config()
    excluded_materials = set(filter_config.get("excluded_materials", []))
    exclude_contract_types = filter_config.get("exclude_contract_types", [])

    print(f"\n📋 开始处理 Excel 文件: {file_path}")
    print(f"   过滤物料数量: {len(excluded_materials)}")
    print(f"   过滤合同类型关键词: {exclude_contract_types}")

    if not Path(file_path).exists():
        print(f"❌ 文件不存在: {file_path}", file=sys.stderr)
        return False

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 找到目标列（按列名动态查找，不依赖固定列号）
        contract_col = None
        material_col = None
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(1, col).value
            if val:
                val_str = str(val).strip()
                if "合同业务类型" in val_str:
                    contract_col = col
                    print(f"   ✓ 找到'合同业务类型'列: 第{col}列")
                elif "物料名称" in val_str:
                    material_col = col
                    print(f"   ✓ 找到'物料名称'列: 第{col}列")

        if contract_col is None and material_col is None:
            print("   ⚠️  未找到目标列，跳过过滤（文件结构可能已变化，请检查列名）")
            workbook.close()
            return True

        total_rows = sheet.max_row - 1
        deleted_contract = 0
        deleted_material = 0

        # 从最后一行向前遍历，避免删行后行号错位
        for row in range(sheet.max_row, 1, -1):
            should_delete = False

            if contract_col and exclude_contract_types:
                val = sheet.cell(row, contract_col).value
                if val:
                    val_str = str(val)
                    if any(kw in val_str for kw in exclude_contract_types):
                        should_delete = True
                        deleted_contract += 1

            if not should_delete and material_col and excluded_materials:
                val = sheet.cell(row, material_col).value
                if val and str(val).strip() in excluded_materials:
                    should_delete = True
                    deleted_material += 1

            if should_delete:
                sheet.delete_rows(row)

        workbook.save(file_path)
        workbook.close()

        total_deleted = deleted_contract + deleted_material
        remaining = total_rows - total_deleted
        print(
            f"✅ Excel 处理完成！原始 {total_rows} 行，"
            f"删除 {total_deleted} 行（合同类型 {deleted_contract}，物料名称 {deleted_material}），"
            f"剩余 {remaining} 行"
        )
        return True

    except Exception as e:
        print(f"❌ Excel 处理失败: {e}", file=sys.stderr)
        import traceback

        traceback.print_exc()
        return False


def main():
    parser = argparse.ArgumentParser(description="过滤 iERP 导出的出库单 Excel")
    parser.add_argument("--file", required=True, help="Excel 文件路径")
    args = parser.parse_args()

    success = filter_excel(args.file)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
