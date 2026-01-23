#!/usr/bin/env python3
"""
Excel 文件读取工具
用于读取《数据标准字典-销项发票数据》V0.1.xlsx 并转换为可读格式
"""

import sys
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def read_excel_file(file_path: str, output_format: str = "markdown") -> str:
    """
    读取 Excel 文件并转换为指定格式
    
    Args:
        file_path: Excel 文件路径
        output_format: 输出格式 (markdown, json, text)
    
    Returns:
        格式化后的内容字符串
    """
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        result = []
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            if output_format == "markdown":
                result.append(f"\n## {sheet_name}\n")
                result.append(format_sheet_as_markdown(sheet))
            elif output_format == "json":
                result.append(format_sheet_as_json(sheet, sheet_name))
            else:  # text
                result.append(f"\n=== {sheet_name} ===\n")
                result.append(format_sheet_as_text(sheet))
        
        if output_format == "json":
            return json.dumps({"sheets": result}, ensure_ascii=False, indent=2)
        else:
            return "\n".join(result)
    
    except FileNotFoundError:
        return f"错误: 文件不存在 - {file_path}"
    except Exception as e:
        return f"错误: 读取文件失败 - {str(e)}"


def format_sheet_as_markdown(sheet) -> str:
    """将工作表格式化为 Markdown 表格"""
    lines = []
    rows = list(sheet.iter_rows(values_only=True))
    
    if not rows:
        return "*空表*\n"
    
    # 获取表头
    headers = rows[0]
    if not any(headers):
        return "*空表*\n"
    
    # 构建表头
    header_line = "| " + " | ".join(str(h or "") for h in headers) + " |"
    separator = "| " + " | ".join("---" for _ in headers) + " |"
    lines.append(header_line)
    lines.append(separator)
    
    # 构建数据行
    for row in rows[1:]:
        if not any(row):  # 跳过空行
            continue
        row_line = "| " + " | ".join(str(cell or "") for cell in row) + " |"
        lines.append(row_line)
    
    return "\n".join(lines) + "\n"


def format_sheet_as_json(sheet, sheet_name: str) -> dict:
    """将工作表格式化为 JSON"""
    rows = list(sheet.iter_rows(values_only=True))
    
    if not rows or not any(rows[0]):
        return {"name": sheet_name, "data": []}
    
    headers = [str(h or f"Column_{i}") for i, h in enumerate(rows[0])]
    data = []
    
    for row in rows[1:]:
        if not any(row):
            continue
        row_dict = {headers[i]: (cell if cell is not None else "") 
                   for i, cell in enumerate(row)}
        data.append(row_dict)
    
    return {"name": sheet_name, "headers": headers, "data": data}


def format_sheet_as_text(sheet) -> str:
    """将工作表格式化为纯文本"""
    lines = []
    rows = list(sheet.iter_rows(values_only=True))
    
    if not rows:
        return "空表\n"
    
    for row in rows:
        if not any(row):
            continue
        line = "\t".join(str(cell or "") for cell in row)
        lines.append(line)
    
    return "\n".join(lines) + "\n"


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python excel_reader.py <excel_file_path> [output_format]")
        print("output_format: markdown (默认), json, text")
        sys.exit(1)
    
    file_path = sys.argv[1]
    output_format = sys.argv[2] if len(sys.argv) > 2 else "markdown"
    
    if output_format not in ["markdown", "json", "text"]:
        print(f"错误: 不支持的输出格式 '{output_format}'")
        print("支持的格式: markdown, json, text")
        sys.exit(1)
    
    result = read_excel_file(file_path, output_format)
    print(result)


if __name__ == "__main__":
    main()
