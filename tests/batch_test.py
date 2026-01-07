#!/usr/bin/env python
"""
批量自动化测试脚本 - 测试 customer-service skill

Usage:
    # 基本用法
    python tests/batch_test.py tests/dataset/test_set_1.md

    # 控制并发数（每个问题约3-4分钟，建议并发数<=3）
    python tests/batch_test.py tests/dataset/test_set_1.md --concurrency 3

    # 指定默认产品（当 agent 询问产品时自动回复）
    python tests/batch_test.py tests/dataset/test_set_1.md --default-product "星瀚旗舰版"

    # 调整超时（默认360秒）
    python tests/batch_test.py tests/dataset/test_set_1.md --timeout 600

Features:
    - 并发测试（不同session）
    - 自动检测产品询问并自动回复默认产品
    - 超时时保留部分回答
    - 输出 Excel (.xlsx) + JSON 结果到 tests/results/
    - Excel 格式支持长文本自动换行，完整内容不截断

测试数据格式：
    每行一个问题，支持 - 或 * 前缀，# 开头为注释

注意事项：
    - 每个问题处理时间约 3-4 分钟
    - 默认超时 360 秒（6分钟）
    - 结果保存在 tests/results/ 目录
"""

import asyncio
import json
import logging
import re
import sys
import argparse
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, asdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# 确保项目根目录在 Python 路径中
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv

# Load environment variables
load_dotenv(PROJECT_ROOT / '.env.prod')

from api.dependencies import get_agent_service
from api.models.requests import QueryRequest

# 配置日志 - 分离文件日志和控制台输出
log_dir = PROJECT_ROOT / "logs"
log_dir.mkdir(exist_ok=True)

# 配置根logger
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)

# 清除默认handlers
root_logger.handlers.clear()

# 文件handler - 记录所有INFO及以上级别的日志
file_handler = logging.FileHandler(
    log_dir / "batch.log",
    encoding='utf-8'
)
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter(
    '%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
file_handler.setFormatter(file_formatter)

# 控制台handler - 只显示WARNING及以上（避免中间日志污染输出）
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.WARNING)
console_formatter = logging.Formatter('%(levelname)s: %(message)s')
console_handler.setFormatter(console_formatter)

# 添加handlers
root_logger.addHandler(file_handler)
root_logger.addHandler(console_handler)
logger = logging.getLogger(__name__)


@dataclass
class TestResult:
    """单个测试结果"""
    question: str
    answer: str = ""  # 完整回答（向后兼容）
    thinking: str = ""  # 思考过程
    final_answer: str = ""  # 最终回答
    session_id: str = ""
    rounds: int = 1  # 对话轮数
    duration_ms: float = 0
    status: str = "pending"  # pending, success, error, needs_product
    error: str = ""
    product_selected: str = ""  # 如果触发了产品选择


# 最终回答的分隔标记（按优先级排序）
FINAL_ANSWER_MARKERS = [
    "## 核心结论",
    "## 最终回答",
    "## 总结",
    "## 结论",
    "**核心结论**",
    "**最终回答**",
    "**总结**",
    "**结论**",
    "根据知识库",
    "根据文档",
]


def split_thinking_and_answer(full_answer: str) -> tuple[str, str]:
    """将完整回答分割为思考过程和最终回答

    Returns:
        (thinking, final_answer) 元组
    """
    if not full_answer:
        return "", ""

    # 尝试按标记分割
    for marker in FINAL_ANSWER_MARKERS:
        if marker in full_answer:
            idx = full_answer.find(marker)
            thinking = full_answer[:idx].strip()
            final_answer = full_answer[idx:].strip()
            return thinking, final_answer

    # 如果没有找到标记，尝试按"---"分隔符分割（多轮对话）
    if "\n---\n" in full_answer:
        parts = full_answer.split("\n---\n")
        if len(parts) >= 2:
            # 最后一部分通常是最终回答
            thinking = "\n---\n".join(parts[:-1])
            final_answer = parts[-1]
            return thinking.strip(), final_answer.strip()

    # 无法分割，全部作为最终回答
    return "", full_answer


# 产品选择检测模式 - 只匹配明确的询问，避免匹配陈述句
PRODUCT_ASK_PATTERNS = [
    r"请问您使用的是哪个产品",
    r"您使用的是.*产品.*[？?]",  # 必须是问句（以问号结尾）
    r"请确认.*产品版本",
    r"请选择.*产品",
    r"1\.\s*标准版.*\n.*2\.\s*星瀚",  # 产品列表选项
]

# 产品关键词到回复的映射（值为发送给Claude的完整回复）
PRODUCT_REPLIES = {
    "标准版发票云": "我使用的是标准版发票云",
    "标准版": "我使用的是标准版发票云",
    "星瀚旗舰版": "我使用的是星瀚旗舰版",
    "星瀚": "我使用的是星瀚旗舰版",
    "旗舰版发票云": "我使用的是星瀚旗舰版",  # 默认指星瀚旗舰版
    "星空旗舰版": "我使用的是星空旗舰版",
    "星空": "我使用的是星空旗舰版",
}


def detect_product_question(text: str) -> bool:
    """检测是否在询问产品选择

    Returns:
        True if the text contains a product question, False otherwise
    """
    # 如果回答已经包含完成标记，说明不是在询问产品
    completion_markers = ["执行进度完成", "## 核心结论", "## 最终回答", "## 总结"]
    for marker in completion_markers:
        if marker in text:
            logger.debug(f"检测到完成标记 '{marker}'，跳过产品询问检测")
            return False

    for pattern in PRODUCT_ASK_PATTERNS:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            logger.info(f"产品询问检测命中: pattern='{pattern}', matched='{match.group()}'")
            return True
    return False


def parse_test_questions(file_path: str) -> list[str]:
    """解析测试问题文件

    支持格式：
    - 每行一个问题
    - 以 - 开头的列表
    - 空行和 # 开头的注释会被跳过
    """
    questions = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            # 跳过空行和注释
            if not line or line.startswith('#'):
                continue
            # 移除列表前缀
            if line.startswith('- '):
                line = line[2:]
            elif line.startswith('* '):
                line = line[2:]
            # 移除序号前缀 (1. 2. 等)
            line = re.sub(r'^\d+\.\s*', '', line)
            if line:
                questions.append(line)
    return questions


class TestRunner:
    """单个测试的运行器，用于保持状态以便超时时恢复"""

    def __init__(self, agent_service, question: str, default_product: str,
                 max_rounds: int = 3, task_id: str = ""):
        self.agent_service = agent_service
        self.question = question
        self.default_product = default_product
        self.max_rounds = max_rounds
        self.task_id = task_id  # 任务标识，如 "[1/10]"

        # 状态
        self.result = TestResult(question=question)
        self.session_id = None
        self.full_answer = []
        self.round_answer = []
        self.start_time = datetime.now()

    def log(self, msg: str, level: str = "info"):
        """带任务标识的日志"""
        full_msg = f"{self.task_id} {msg}" if self.task_id else msg
        getattr(logger, level)(full_msg)

    def finalize_result(self, status: str, error: str = "", partial: bool = False):
        """整理最终结果"""
        # 收集所有回答
        if self.round_answer:
            self.full_answer.append("".join(self.round_answer))

        if self.full_answer:
            answer = "\n---\n".join(self.full_answer)
            if partial:
                answer += f"\n[{status.upper()} - 回答不完整]"
            self.result.answer = answer
            # 分离思考过程和最终回答
            thinking, final_answer = split_thinking_and_answer(answer)
            self.result.thinking = thinking
            self.result.final_answer = final_answer

        self.result.status = status
        self.result.error = error
        self.result.duration_ms = (datetime.now() - self.start_time).total_seconds() * 1000

        # 记录最终状态
        duration_s = self.result.duration_ms / 1000
        if status == "success":
            self.log(f"测试成功 ({duration_s:.1f}s)")
        elif status == "timeout":
            self.log(f"测试超时 ({duration_s:.1f}s)", "warning")
        elif status == "error":
            self.log(f"测试失败: {error} ({duration_s:.1f}s)", "error")
        else:
            self.log(f"测试结束: {status} ({duration_s:.1f}s)")

        return self.result

    async def run(self) -> TestResult:
        """运行测试"""
        try:
            for round_num in range(1, self.max_rounds + 1):
                self.result.rounds = round_num
                self.round_answer = []
                self.log(f"第{round_num}轮开始")

                # 构建请求
                if round_num == 1:
                    prompt = self.question
                else:
                    prompt = PRODUCT_REPLIES.get(self.default_product, f"我使用的是{self.default_product}")
                    self.result.product_selected = self.default_product
                    self.log(f"第{round_num}轮 - 回复产品: {prompt}")

                request = QueryRequest(
                    tenant_id="batch-test",
                    prompt=prompt,
                    skill="customer-service",
                    language="中文" if not self.session_id else None,
                    session_id=self.session_id,
                    metadata={"source": "batch-test"}
                )

                # 处理流式响应
                async for message in self.agent_service.process_query(request):
                    event_type = message.get("event")
                    data = message.get("data")

                    if event_type == "heartbeat":
                        continue

                    try:
                        data_obj = json.loads(data) if isinstance(data, str) else data
                    except json.JSONDecodeError:
                        data_obj = {"raw": data}

                    if event_type == "session_created":
                        self.session_id = data_obj.get("session_id")
                        self.result.session_id = self.session_id or ""
                        self.log(f"会话创建: {self.session_id[:16]}..." if self.session_id else "会话创建失败")

                    elif event_type == "assistant_message":
                        content = data_obj.get("content", "")
                        if content:
                            self.round_answer.append(content)

                    elif event_type == "result":
                        self.result.duration_ms = data_obj.get("duration_ms", 0)

                    elif event_type == "error":
                        self.log(f"收到错误: {data_obj.get('message', str(data_obj))}", "error")
                        return self.finalize_result("error", data_obj.get("message", str(data_obj)), partial=True)

                # 完成一轮
                round_text = "".join(self.round_answer)
                self.full_answer.append(round_text)
                self.round_answer = []
                self.log(f"第{round_num}轮完成, 回答长度: {len(round_text)}")

                # 检测是否需要产品选择
                if detect_product_question(round_text):
                    if round_num < self.max_rounds and self.default_product:
                        reply_text = PRODUCT_REPLIES.get(self.default_product, f"我使用的是{self.default_product}")
                        auto_reply_note = f"\n\n[批量测试] 检测到产品询问，自动回复: {reply_text}\n"
                        self.full_answer.append(auto_reply_note)
                        self.log(f"检测到产品询问，自动回复: {reply_text}")
                        continue
                    else:
                        return self.finalize_result("needs_product")
                else:
                    break

            return self.finalize_result("success")

        except asyncio.CancelledError:
            self.log("任务被取消", "warning")
            return self.finalize_result("timeout", "Task cancelled", partial=True)
        except Exception as e:
            self.log(f"异常: {e}", "error")
            logger.exception(f"{self.task_id} Test failed for: {self.question[:50]}...")
            return self.finalize_result("error", str(e), partial=True)


async def run_single_test(
    agent_service,
    question: str,
    default_product: str,
    max_rounds: int = 3
) -> TestRunner:
    """创建并返回 TestRunner（不等待完成）"""
    runner = TestRunner(agent_service, question, default_product, max_rounds)
    return runner


async def run_batch_tests(
    questions: list[str],
    concurrency: int = 5,
    default_product: str = "旗舰版发票云",
    timeout: float = 300.0
) -> list[TestResult]:
    """并发运行批量测试"""
    agent_service = get_agent_service()
    semaphore = asyncio.Semaphore(concurrency)

    def log_progress(msg: str):
        """同时输出到控制台和日志文件"""
        print(msg)
        logger.info(msg)

    async def run_with_semaphore(idx: int, question: str) -> tuple[int, TestResult]:
        task_id = f"[{idx+1}/{len(questions)}]"
        async with semaphore:
            log_progress(f"{task_id} 开始: {question[:40]}...")
            runner = TestRunner(agent_service, question, default_product, task_id=task_id)

            try:
                # 使用 wait_for 但保持 runner 引用
                result = await asyncio.wait_for(runner.run(), timeout=timeout)
                status_icon = "✓" if result.status == "success" else ("⏱" if result.status == "timeout" else "✗")
                log_progress(f"{task_id} {status_icon} 完成 ({result.duration_ms/1000:.1f}s, {result.rounds}轮)")
                return idx, result
            except asyncio.TimeoutError:
                # 超时时，从 runner 中提取部分结果
                log_progress(f"{task_id} ⏱ 超时 ({timeout}s)")
                result = runner.finalize_result("timeout", f"Timeout after {timeout}s", partial=True)
                return idx, result
            except asyncio.CancelledError:
                log_progress(f"{task_id} ⏱ 取消")
                result = runner.finalize_result("timeout", "Task cancelled", partial=True)
                return idx, result
            except Exception as e:
                log_progress(f"{task_id} ✗ 异常: {e}")
                logger.exception(f"{task_id} exception")
                result = runner.finalize_result("error", str(e), partial=True)
                return idx, result

    # 创建所有任务
    tasks = [
        run_with_semaphore(idx, q)
        for idx, q in enumerate(questions)
    ]

    # 并发执行
    completed = await asyncio.gather(*tasks, return_exceptions=True)

    # 按原始顺序整理结果
    results = [None] * len(questions)
    for item in completed:
        if isinstance(item, Exception):
            logger.error(f"Gather exception: {item}")
            continue
        idx, result = item
        results[idx] = result

    # 填充失败的结果
    for idx, r in enumerate(results):
        if r is None:
            results[idx] = TestResult(
                question=questions[idx],
                status="error",
                error="Task failed unexpectedly"
            )

    return results


def save_results_xlsx(results: list[TestResult], output_dir: Path, name: str) -> Path:
    """保存测试结果到 Excel (.xlsx) 格式 - 完整内容，自动换行"""
    if not XLSX_AVAILABLE:
        print("⚠️  openpyxl 未安装，跳过 xlsx 输出。安装命令: pip install openpyxl")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"{name}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "测试结果"

    # 定义列宽（字符数）
    column_widths = {
        'A': 8,   # 序号
        'B': 40,  # 问题
        'C': 80,  # 最终回答
        'D': 80,  # 思考过程
        'E': 12,  # 状态
        'F': 12,  # 耗时
        'G': 8,   # 轮数
        'H': 15,  # 选择的产品
        'I': 30   # 错误信息
    }

    # 设置列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ["序号", "问题", "最终回答", "思考过程", "状态", "耗时(s)", "轮数", "选择的产品", "错误信息"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 内容样式
    content_alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True  # 自动换行
    )

    # 写入数据
    for idx, r in enumerate(results, 1):
        row_idx = idx + 1

        # 最终回答优先显示，如果为空则显示完整回答
        final = r.final_answer if r.final_answer else r.answer

        # 写入各列数据
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=r.question)
        ws.cell(row=row_idx, column=3, value=final)  # 完整内容，不截断
        ws.cell(row=row_idx, column=4, value=r.thinking)  # 完整内容，不截断
        ws.cell(row=row_idx, column=5, value=r.status)
        ws.cell(row=row_idx, column=6, value=f"{r.duration_ms/1000:.1f}")
        ws.cell(row=row_idx, column=7, value=r.rounds)
        ws.cell(row=row_idx, column=8, value=r.product_selected)
        ws.cell(row=row_idx, column=9, value=r.error)

        # 应用样式到所有内容单元格
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = content_alignment

        # 状态列居中
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="center")

        # 根据状态设置颜色
        status_cell = ws.cell(row=row_idx, column=5)
        if r.status == "success":
            status_cell.font = Font(color="008000")  # 绿色
        elif r.status == "timeout":
            status_cell.font = Font(color="FF8C00")  # 橙色
        elif r.status == "error":
            status_cell.font = Font(color="FF0000")  # 红色

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(xlsx_path)
    print(f"✅ Excel 结果已保存: {xlsx_path}")
    logger.info(f"Excel 结果已保存: {xlsx_path}")

    return xlsx_path


def save_results(results: list[TestResult], output_dir: Path, name: str):
    """保存测试结果到 JSON 和 Excel"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{name}_{timestamp}"

    # 保存 JSON
    json_path = output_dir / f"{base_name}.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)
    print(f"JSON 结果已保存: {json_path}")
    logger.info(f"JSON 结果已保存: {json_path}")

    # 保存 Excel (支持长文本和自动换行)
    save_results_xlsx(results, output_dir, name)

    # 打印摘要
    success = sum(1 for r in results if r.status == "success")
    timeout = sum(1 for r in results if r.status == "timeout")
    errors = sum(1 for r in results if r.status == "error")
    needs_product = sum(1 for r in results if r.status == "needs_product")

    print(f"\n{'='*50}")
    print(f"测试完成: {len(results)} 个问题")
    print(f"  ✓ 成功: {success}")
    print(f"  ⏱ 超时: {timeout}")
    print(f"  ✗ 错误: {errors}")
    print(f"  ? 需要产品选择: {needs_product}")
    print(f"{'='*50}\n")

    return json_path


async def main():
    parser = argparse.ArgumentParser(description="批量测试 customer-service agent")
    parser.add_argument("input_file", help="测试问题文件路径")
    parser.add_argument("--concurrency", "-c", type=int, default=5, help="并发数 (默认: 5)")
    parser.add_argument("--default-product", "-p", default="旗舰版发票云",
                        help="默认产品选择 (默认: 旗舰版发票云)")
    parser.add_argument("--timeout", "-t", type=float, default=360.0, help="单个测试超时(秒)，默认360秒")
    parser.add_argument("--output-dir", "-o", default="tests/results", help="输出目录")

    args = parser.parse_args()

    # 解析测试问题
    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"错误: 文件不存在 {input_path}")
        sys.exit(1)

    questions = parse_test_questions(str(input_path))
    if not questions:
        print("错误: 未找到有效的测试问题")
        sys.exit(1)

    print(f"加载了 {len(questions)} 个测试问题")
    print(f"并发数: {args.concurrency}")
    print(f"默认产品: {args.default_product}")
    print()

    # 创建输出目录
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 运行测试
    results = await run_batch_tests(
        questions,
        concurrency=args.concurrency,
        default_product=args.default_product,
        timeout=args.timeout
    )

    # 保存结果
    save_results(results, output_dir, input_path.stem)


if __name__ == "__main__":
    asyncio.run(main())
